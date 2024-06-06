import telebot
from telebot import types
import openpyxl
import os
import matplotlib.pyplot as plt

TOKEN = "7046928663:AAHvULiD3cI4Y-_y-xvXmdKfN-r-LPP1M-o"
EXCEL_FOLDER = 'excel_files'
INCOME_SOURCES = ['Salary', 'Family', 'Gift', 'Refund', 'Sales']
EXPENSE_CATEGORIES = ['Health', 'Clothing', 'Food', 'Housing', 'Transportation', 'Entertainment', 'Other']

bot = telebot.TeleBot(TOKEN)

if not os.path.exists(EXCEL_FOLDER):
    os.makedirs(EXCEL_FOLDER)   


def create_main_menu():
    markup = types.InlineKeyboardMarkup(row_width=2)
    expense_button = types.InlineKeyboardButton('Траты', callback_data='expense')
    income_button = types.InlineKeyboardButton('Доходы', callback_data='income')
    markup.add(expense_button, income_button)
    return markup


def create_income_menu():
    markup = types.InlineKeyboardMarkup(row_width=2)
    for source in INCOME_SOURCES:
        button = types.InlineKeyboardButton(source, callback_data=f'income_{source.lower()}')
        markup.add(button)
    back_button = types.InlineKeyboardButton('Назад', callback_data='back')
    markup.add(back_button)
    return markup


def create_expense_menu():
    markup = types.InlineKeyboardMarkup(row_width=2)
    for category in EXPENSE_CATEGORIES:
        button = types.InlineKeyboardButton(category, callback_data=f'expense_{category.lower()}')
        markup.add(button)
    back_button = types.InlineKeyboardButton('Назад', callback_data='back')
    markup.add(back_button)
    return markup


@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(message.chat.id, "Привет! Я бот для отслеживания трат и доходов.")


@bot.message_handler(commands=['new'])
def new(message):
    bot.send_message(message.chat.id, "Выберите тип операции:", reply_markup=create_main_menu())


@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    if call.message:
        if call.data == 'expense':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text="Выберите категорию трат:", reply_markup=create_expense_menu())
        elif call.data == 'income':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text="Выберите источник дохода:", reply_markup=create_income_menu())
        elif call.data.startswith('income_'):
            category = call.data.split('_')[1]
            operation_type = 'Income'
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=f"Введите сумму и описание дохода в категории {category}:")
            bot.register_next_step_handler(call.message,
                                           lambda message: handle_message(message, operation_type, category))
        elif call.data.startswith('expense_'):
            category = call.data.split('_')[1]
            operation_type = 'Expense'
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=f"Введите сумму и описание расхода в категории {category}:")
            bot.register_next_step_handler(call.message,
                                           lambda message: handle_message(message, operation_type, category))
        elif call.data == 'back':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text="Выберите тип операции:", reply_markup=create_main_menu())
        else:
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text="Ошибка в команде.")


def handle_message(message, operation_type, category):
    chat_id = message.chat.id
    text = message.text

    try:
        amount_str, description = text.split(maxsplit=1)
        amount = float(amount_str)
        row = [operation_type, category, amount, description]
        save_record(chat_id, row)
        bot.send_message(chat_id, "Запись успешно сохранена в Excel файле.")
    except ValueError:
        bot.send_message(chat_id, "Некорректный формат сообщения. Пожалуйста попробуйте создать новую запись заново")


def save_record(chat_id, row):
    excel_filename = f"{chat_id}_expenses.xlsx"
    excel_path = os.path.join(EXCEL_FOLDER, excel_filename)
    if not os.path.exists(excel_path):
        workbook = openpyxl.Workbook()
        workbook.save(excel_path)

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    sheet.append(row)
    workbook.save(excel_path)


def send_statistics(chat_id, title, labels, values):
    plt.figure(figsize=(8, 5))
    plt.pie(values, labels=labels, autopct='%1.1f%%')
    plt.axis('equal')
    plt.title(title)

    plt.savefig('stats.png')
    bot.send_photo(chat_id, open('stats.png', 'rb'))
    plt.close()


@bot.message_handler(commands=['stats'])
def show_statistics(message):
    chat_id = message.chat.id
    excel_filename = f"{chat_id}_expenses.xlsx"
    excel_path = os.path.join(EXCEL_FOLDER, excel_filename)

    if not os.path.exists(excel_path):
        bot.send_message(chat_id, "У вас пока нет данных о расходах и доходах.")
        return

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    total_income = sum(row[2].value for row in sheet.iter_rows(min_row=2) if row[0].value == 'Income')
    total_expense = sum(row[2].value for row in sheet.iter_rows(min_row=2) if row[0].value == 'Expense')
    balance = total_income - total_expense

    message = f"Текущий баланс: {balance}\nОбщие траты: {total_expense}\nОбщие доходы: {total_income}"
    bot.send_message(chat_id, message)

    income_categories = {row[1].value: row[2].value for row in sheet.iter_rows(min_row=2) if row[0].value == 'Income'}
    send_statistics(chat_id, 'Статистика по доходам', list(income_categories.keys()), list(income_categories.values()))

    expense_categories = {row[1].value: row[2].value for row in sheet.iter_rows(min_row=2) if row[0].value == 'Expense'}
    send_statistics(chat_id, 'Статистика по расходам', list(expense_categories.keys()),
                    list(expense_categories.values()))



bot.polling()
